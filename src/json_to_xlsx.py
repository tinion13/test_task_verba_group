import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _extract_country(characteristics):
    for ch in characteristics:
        if ch.get("name") == "Страна производства":
            return ch.get("value")
    return None


def _extract_price_and_stock(sizes):
    size_names = []
    prices_raw = []
    total_stock = 0

    for size in sizes:
        name = size.get("name")
        price = size.get("price")
        remains = size.get("remains")

        if name is not None:
            size_names.append(str(name))

        if price is not None:
            prices_raw.append(price)

        if isinstance(remains, (int, float)):
            total_stock += remains

    price_rub = min(prices_raw) / 100 if prices_raw else None
    sizes_str = ", ".join(size_names)

    return price_rub, total_stock, sizes_str


def _build_rows(data):
    rows = []
    for item in data:
        characteristics = item.get("characteristics", [])
        sizes = item.get("sizes", [])
        price_rub, total_stock, sizes_str = _extract_price_and_stock(sizes)
        country = _extract_country(characteristics)
        rows.append({
            "Ссылка на товар": item.get("url"),
            "Артикул": item.get("article"),
            "Название": item.get("name"),
            "Цена": price_rub,
            "Описание": item.get("description"),
            "Ссылки на изображения через запятую": item.get("images_urls"),
            "Все характеристики с сохранением их структуры": json.dumps(
                characteristics, ensure_ascii=False
            ),
            "Название селлера": item.get("seller_name"),
            "Ссылка на селлера": item.get("seller_url"),
            "Размеры товара через запятую": sizes_str,
            "Остатки по товару (число)": total_stock,
            "Рейтинг": item.get("rating"),
            "Количество отзывов": item.get("number_reviews"),
            "Страна производства": country,
        })
    return rows


def _write_xlsx(rows, xlsx_filename, include_country=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"

    headers = [
        "Ссылка на товар",
        "Артикул",
        "Название",
        "Цена",
        "Описание",
        "Ссылки на изображения через запятую",
        "Все характеристики с сохранением их структуры",
        "Название селлера",
        "Ссылка на селлера",
        "Размеры товара через запятую",
        "Остатки по товару (число)",
        "Рейтинг",
        "Количество отзывов",
    ]

    if include_country:
        headers.append("Страна производства")

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(h) for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)

        if header in ("Описание", "Все характеристики с сохранением их структуры",
                      "Ссылки на изображения через запятую"):
            width = 45
        elif header in ("Ссылка на товар", "Ссылка на селлера", "Название"):
            width = 28
        else:
            width = 18

        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if "Цена" in headers:
        price_col = headers.index("Цена") + 1
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=price_col).number_format = '#,##0.00'

    wb.save(xlsx_filename)


def convert_json_to_xlsx_with_filtered(
    json_filename: str,
    xlsx_filename: str,
    xlsx_filename_filtered: str
):
    with open(json_filename, encoding="utf-8") as f:
        data = json.load(f)

    rows = _build_rows(data)

    _write_xlsx(rows, xlsx_filename, include_country=True)

    filtered_rows = []
    for row in rows:
        rating = row.get("Рейтинг")
        price = row.get("Цена")
        country = row.get("Страна производства")

        if (
            rating is not None and rating >= 4.5
            and price is not None and price <= 10000
            and country == "Россия"
        ):
            filtered_rows.append(row)

    _write_xlsx(filtered_rows, xlsx_filename_filtered, include_country=True)